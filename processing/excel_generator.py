from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def create_excel(all_documents, output_path):

    wb = Workbook()

    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # =====================================================
    # One metadata sheet for each uploaded file
    # =====================================================
    for doc_index, document in enumerate(all_documents, start=1):

        metadata = document["metadata"]

        metadata_sheet = wb.create_sheet(
            f"Metadata_{doc_index}"
        )

        metadata_sheet["A1"] = "Field"
        metadata_sheet["B1"] = "Value"

        metadata_sheet["A1"].font = Font(bold=True)
        metadata_sheet["B1"].font = Font(bold=True)

        row = 2

        for key, value in metadata.items():

            metadata_sheet.cell(row=row, column=1).value = key
            metadata_sheet.cell(row=row, column=2).value = value

            row += 1

    # =====================================================
    # Data Sheets
    # =====================================================
    for doc_index, document in enumerate(all_documents, start=1):

        dfs = document["dataframes"]

        filename = document["filename"]

        sheet_names = document["sheet_names"]

        base_name = filename.rsplit(".", 1)[0]

        for i, df in enumerate(dfs):

            if sheet_names and i < len(sheet_names):

                sheet_title = f"{base_name}_{sheet_names[i]}"

            else:

                sheet_title = f"{base_name}_Table_{i+1}"

            # Excel sheet name max = 31 chars
            sheet_title = sheet_title[:31]

            ws = wb.create_sheet(sheet_title)

            if df.empty:
                continue

            # Header
            for col, header in enumerate(df.columns, start=1):

                cell = ws.cell(row=1, column=col)

                cell.value = str(header)

                cell.font = Font(bold=True)

            # Data
            for r, row in enumerate(
                df.itertuples(index=False),
                start=2
            ):

                for c, value in enumerate(row, start=1):

                    ws.cell(
                        row=r,
                        column=c
                    ).value = value

            ws.freeze_panes = "A2"

            # Auto width
            for column_cells in ws.columns:

                length = max(
                    len(str(cell.value))
                    if cell.value is not None
                    else 0
                    for cell in column_cells
                )

                ws.column_dimensions[
                    get_column_letter(
                        column_cells[0].column
                    )
                ].width = min(length + 3, 50)

    wb.save(output_path)